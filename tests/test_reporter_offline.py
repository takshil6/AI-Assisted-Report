"""
Offline tests for Phase 4 reporter. Uses a minimal analysis JSON fixture
so the tests don't require Phase 2 or Phase 3 to have run.

Run:
    python tests/test_reporter_offline.py
"""

import json
import sys
import tempfile
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from openpyxl import load_workbook

from reporter import generate_report


def _fixture_analysis() -> dict:
    """Minimal analysis dict with the shape Phase 3 produces."""
    return {
        "dataset_summary": {
            "total_revenue":     1000.0,
            "total_transactions": 10,
            "date_range_start":   "2024-01-01",
            "date_range_end":     "2024-01-28",
            "unique_customers":   3,
            "unique_segments":    2,
            "unique_regions":     2,
        },
        "pareto_summary": {
            "total_customers":      3,
            "pareto_customers":     1,
            "pareto_customer_pct":  33.3,
            "pareto_revenue_pct":   80.0,
            "top_1_customer":       "Customer A",
            "top_1_revenue_pct":    80.0,
        },
        "weekly_summary": {
            "weeks_covered":        4,
            "avg_weekly_revenue":   250.0,
            "best_week_revenue":    300.0,
            "worst_week_revenue":   200.0,
            "latest_week_revenue":  250.0,
            "latest_wow_change":    0.0,
        },
        "top_customers": [
            {"rank": 1, "customer_name_clean": "Customer A",
             "revenue": 800.0, "transactions": 4, "avg_order_value": 200.0,
             "pct_of_total": 80.0, "cumulative_pct": 80.0, "is_pareto_80": True},
            {"rank": 2, "customer_name_clean": "Customer B",
             "revenue": 150.0, "transactions": 3, "avg_order_value": 50.0,
             "pct_of_total": 15.0, "cumulative_pct": 95.0, "is_pareto_80": False},
            {"rank": 3, "customer_name_clean": "Customer C",
             "revenue": 50.0, "transactions": 1, "avg_order_value": 50.0,
             "pct_of_total": 5.0, "cumulative_pct": 100.0, "is_pareto_80": False},
        ],
        "segment_breakdown": [
            {"industry_segment_clean": "Paper", "revenue": 700.0,
             "transactions": 7, "unique_customers": 2, "pct_of_total": 70.0},
            {"industry_segment_clean": "Mining", "revenue": 300.0,
             "transactions": 3, "unique_customers": 1, "pct_of_total": 30.0},
        ],
        "region_breakdown": [
            {"region_clean": "East", "revenue": 600.0,
             "transactions": 6, "unique_customers": 2, "pct_of_total": 60.0},
            {"region_clean": "West", "revenue": 400.0,
             "transactions": 4, "unique_customers": 2, "pct_of_total": 40.0},
        ],
    }


def _fixture_weekly() -> pd.DataFrame:
    return pd.DataFrame({
        "order_date_clean": pd.to_datetime(["2024-01-07", "2024-01-14", "2024-01-21", "2024-01-28"]),
        "revenue":          [250.0, 300.0, 250.0, 200.0],
        "transactions":     [3, 3, 2, 2],
        "wow_change_pct":   [None, 20.0, -16.67, -20.0],
    })


def test_generates_six_sheets():
    with tempfile.TemporaryDirectory() as tmp:
        analysis_path = Path(tmp) / "analysis.json"
        weekly_path   = Path(tmp) / "weekly.csv"
        analysis_path.write_text(json.dumps(_fixture_analysis()))
        _fixture_weekly().to_csv(weekly_path, index=False)

        out = generate_report(
            analysis_path=str(analysis_path),
            weekly_path=str(weekly_path),
            output_dir=tmp,
        )

        wb = load_workbook(out)
        assert wb.sheetnames == [
            "Executive Summary", "80-20 Customers", "Segments",
            "Regions", "Weekly Trend", "Pareto Curve",
        ], f"Got: {wb.sheetnames}"
    print("[PASS] test_generates_six_sheets passed")


def test_vip_rows_highlighted():
    with tempfile.TemporaryDirectory() as tmp:
        analysis_path = Path(tmp) / "analysis.json"
        weekly_path   = Path(tmp) / "weekly.csv"
        analysis_path.write_text(json.dumps(_fixture_analysis()))
        _fixture_weekly().to_csv(weekly_path, index=False)

        out = generate_report(
            analysis_path=str(analysis_path),
            weekly_path=str(weekly_path),
            output_dir=tmp,
        )

        wb = load_workbook(out)
        ws = wb["80-20 Customers"]

        # Row 5 should be Customer A (VIP, green fill)
        vip_fill = ws.cell(5, 2).fill.fgColor.rgb
        vip_bold = ws.cell(5, 2).font.bold
        assert vip_fill == "00D6F0DC", f"VIP row should be green, got {vip_fill}"
        assert vip_bold, "VIP row should be bold"

        # Row 6 should be Customer B (not VIP)
        non_vip_bold = ws.cell(6, 2).font.bold
        assert not non_vip_bold, "Non-VIP row should NOT be bold"

    print("[PASS] test_vip_rows_highlighted passed")


def test_no_formula_errors():
    """
    Sanity check: the file opens in openpyxl without raising. This catches
    the common bug where openpyxl produces a corrupt .xlsx (e.g., bad chart
    references, merged-cell conflicts).
    """
    with tempfile.TemporaryDirectory() as tmp:
        analysis_path = Path(tmp) / "analysis.json"
        weekly_path   = Path(tmp) / "weekly.csv"
        analysis_path.write_text(json.dumps(_fixture_analysis()))
        _fixture_weekly().to_csv(weekly_path, index=False)

        out = generate_report(
            analysis_path=str(analysis_path),
            weekly_path=str(weekly_path),
            output_dir=tmp,
        )

        # If this loads without raising, the file is structurally valid
        wb = load_workbook(out)
        assert wb is not None
    print("[PASS] test_no_formula_errors passed")


def test_charts_embedded():
    with tempfile.TemporaryDirectory() as tmp:
        analysis_path = Path(tmp) / "analysis.json"
        weekly_path   = Path(tmp) / "weekly.csv"
        analysis_path.write_text(json.dumps(_fixture_analysis()))
        _fixture_weekly().to_csv(weekly_path, index=False)

        out = generate_report(
            analysis_path=str(analysis_path),
            weekly_path=str(weekly_path),
            output_dir=tmp,
        )

        wb = load_workbook(out)
        total_charts = sum(len(wb[n]._charts) for n in wb.sheetnames)
        assert total_charts >= 4, f"Expected >=4 charts, got {total_charts}"
    print("[PASS] test_charts_embedded passed")


if __name__ == "__main__":
    test_generates_six_sheets()
    test_vip_rows_highlighted()
    test_no_formula_errors()
    test_charts_embedded()
    print("\nAll Phase 4 offline tests passed. Safe to run the real reporter.")
