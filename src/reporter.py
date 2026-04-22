"""
Phase 4: Weekly Report Generator
=================================
Reads data/analysis_results.json (from Phase 3) and produces a professional
multi-sheet Excel report with embedded charts:

    Sheet 1: Executive Summary   -- headline KPIs, at-a-glance dashboard
    Sheet 2: 80/20 Customers     -- Pareto table with VIP tier highlighted
    Sheet 3: Segment Breakdown   -- revenue by industry segment + bar chart
    Sheet 4: Region Breakdown    -- revenue by region + bar chart
    Sheet 5: Weekly Trend        -- revenue per week + line chart
    Sheet 6: Pareto Curve        -- classic 80/20 S-curve chart

The output filename is timestamped with the latest data-date so weekly runs
don't overwrite each other:
    reports/weekly_report_YYYY-MM-DD.xlsx

Run:
    python src/reporter.py
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Styling constants -- single source of truth so the whole report is cohesive
# ---------------------------------------------------------------------------
FONT_NAME = "Calibri"

TITLE_FONT      = Font(name=FONT_NAME, size=18, bold=True, color="FFFFFF")
SUBTITLE_FONT   = Font(name=FONT_NAME, size=11, italic=True, color="7F7F7F")
HEADER_FONT     = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
KPI_LABEL_FONT  = Font(name=FONT_NAME, size=10, color="7F7F7F")
KPI_VALUE_FONT  = Font(name=FONT_NAME, size=20, bold=True, color="1F2A44")
BODY_FONT       = Font(name=FONT_NAME, size=11)
VIP_FONT        = Font(name=FONT_NAME, size=11, bold=True, color="0B4F2A")

TITLE_FILL    = PatternFill("solid", fgColor="1F2A44")   # dark navy
HEADER_FILL   = PatternFill("solid", fgColor="4A5D7E")   # muted navy
KPI_FILL      = PatternFill("solid", fgColor="F4F6FA")   # soft gray-blue
VIP_FILL      = PatternFill("solid", fgColor="D6F0DC")   # soft green
ALT_ROW_FILL  = PatternFill("solid", fgColor="F9FAFC")   # zebra striping

THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center", indent=1)
RIGHT  = Alignment(horizontal="right",  vertical="center", indent=1)

# Number formats (per xlsx skill conventions)
FMT_CURRENCY = '"$"#,##0;("$"#,##0);"-"'
FMT_PERCENT  = "0.0%"
FMT_INT      = "#,##0"
FMT_DATE     = "yyyy-mm-dd"


# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------
def load_analysis(path: str = "data/analysis_results.json") -> dict:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(
            f"{path} not found. Run Phase 3 first:  python src/analyzer.py"
        )
    return json.loads(p.read_text())


def load_weekly(path: str = "data/weekly_revenue.csv") -> pd.DataFrame:
    return pd.read_csv(path, parse_dates=["order_date_clean"])


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------
def _apply_title(ws, row: int, text: str, subtitle: str, span_cols: int = 6) -> int:
    """Put a banner title and subtitle at the given row. Returns the next free row."""
    ws.cell(row=row, column=1, value=text).font = TITLE_FONT
    ws.cell(row=row, column=1).fill = TITLE_FILL
    ws.cell(row=row, column=1).alignment = LEFT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    ws.row_dimensions[row].height = 36

    ws.cell(row=row + 1, column=1, value=subtitle).font = SUBTITLE_FONT
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=span_cols)
    return row + 3


def _format_header_row(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 24


def build_exec_summary(wb: Workbook, results: dict) -> None:
    """Sheet 1: top-level KPIs and a narrative takeaway."""
    ws = wb.active
    ws.title = "Executive Summary"

    ds = results["dataset_summary"]
    ps = results["pareto_summary"]
    ws_ = results["weekly_summary"]

    next_row = _apply_title(
        ws, 1,
        "Weekly 80/20 Analytics Report",
        f"Data period: {ds['date_range_start']}  -  {ds['date_range_end']}   "
        f"|   Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    )

    # --- KPI tiles (2 rows x 3 cols) ----------------------------------------
    kpis = [
        ("TOTAL REVENUE",        f"${ds['total_revenue']:,.0f}",            FMT_CURRENCY),
        ("TRANSACTIONS",         f"{ds['total_transactions']:,}",           FMT_INT),
        ("UNIQUE CUSTOMERS",     f"{ds['unique_customers']}",               FMT_INT),
        ("80/20 TIER SIZE",      f"{ps['pareto_customers']} customers",     None),
        ("REVENUE CONCENTRATION", f"{ps['pareto_revenue_pct']}% of revenue", None),
        ("AVG WEEKLY REVENUE",   f"${ws_['avg_weekly_revenue']:,.0f}",      FMT_CURRENCY),
    ]

    tile_start_row = next_row
    for i, (label, value, _fmt) in enumerate(kpis):
        r = tile_start_row + (i // 3) * 4
        c = 1 + (i % 3) * 2

        # Label cell
        lbl = ws.cell(row=r, column=c, value=label)
        lbl.font = KPI_LABEL_FONT
        lbl.fill = KPI_FILL
        lbl.alignment = LEFT
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)

        # Value cell
        val = ws.cell(row=r + 1, column=c, value=value)
        val.font = KPI_VALUE_FONT
        val.fill = KPI_FILL
        val.alignment = LEFT
        ws.merge_cells(start_row=r + 1, start_column=c, end_row=r + 2, end_column=c + 1)
        ws.row_dimensions[r + 1].height = 30
        ws.row_dimensions[r + 2].height = 10

    # --- Narrative takeaway -------------------------------------------------
    narrative_row = tile_start_row + 9
    ws.cell(row=narrative_row, column=1, value="Key Takeaway").font = Font(
        name=FONT_NAME, size=13, bold=True, color="1F2A44"
    )
    takeaway = (
        f"The top {ps['pareto_customers']} customer(s) "
        f"({ps['pareto_customer_pct']}% of all customers) generate "
        f"{ps['pareto_revenue_pct']}% of total revenue. "
        f"The #1 customer alone, {ps['top_1_customer']}, represents "
        f"{ps['top_1_revenue_pct']:.1f}% of revenue -- indicating significant "
        f"concentration risk and a clear priority list for account management."
    )
    cell = ws.cell(row=narrative_row + 1, column=1, value=takeaway)
    cell.font = BODY_FONT
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ws.merge_cells(start_row=narrative_row + 1, start_column=1,
                   end_row=narrative_row + 3, end_column=6)
    ws.row_dimensions[narrative_row + 1].height = 50

    # Column widths
    for col_letter, width in zip("ABCDEF", [22, 22, 22, 22, 22, 22]):
        ws.column_dimensions[col_letter].width = width


def build_pareto_sheet(wb: Workbook, results: dict) -> None:
    """Sheet 2: 80/20 customer list with VIP tier highlighted."""
    ws = wb.create_sheet("80-20 Customers")

    next_row = _apply_title(
        ws, 1,
        "Top Customers - Pareto Analysis",
        "Customers whose cumulative revenue reaches 80% are flagged as the 80/20 tier.",
        span_cols=7,
    )

    headers = ["Rank", "Customer", "Revenue", "Transactions", "Avg Order", "% of Total", "Cumulative %"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=next_row, column=i, value=h)
    _format_header_row(ws, next_row, len(headers))

    data_start = next_row + 1
    customers = results["top_customers"]
    for i, row in enumerate(customers):
        r = data_start + i
        is_vip = row["is_pareto_80"]

        ws.cell(row=r, column=1, value=row["rank"]).alignment = CENTER
        ws.cell(row=r, column=2, value=row["customer_name_clean"]).alignment = LEFT
        ws.cell(row=r, column=3, value=row["revenue"]).number_format = FMT_CURRENCY
        ws.cell(row=r, column=4, value=row["transactions"]).number_format = FMT_INT
        ws.cell(row=r, column=5, value=row["avg_order_value"]).number_format = FMT_CURRENCY
        ws.cell(row=r, column=6, value=row["pct_of_total"] / 100).number_format = FMT_PERCENT
        ws.cell(row=r, column=7, value=row["cumulative_pct"] / 100).number_format = FMT_PERCENT

        # VIP tier styling
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            if is_vip:
                cell.fill = VIP_FILL
                cell.font = VIP_FONT
            else:
                cell.font = BODY_FONT
                if i % 2 == 1:
                    cell.fill = ALT_ROW_FILL

    # Column widths
    for col, w in zip("ABCDEFG", [8, 32, 16, 14, 14, 12, 14]):
        ws.column_dimensions[col].width = w

    ws.freeze_panes = ws.cell(row=data_start, column=1)


def _build_breakdown_sheet(
    wb: Workbook, title: str, sheet_name: str, subtitle: str,
    items: list[dict], key: str, chart_title: str,
) -> None:
    """Shared builder for segment + region sheets since they're identical in structure."""
    ws = wb.create_sheet(sheet_name)

    next_row = _apply_title(ws, 1, title, subtitle, span_cols=5)

    headers = [key.replace("_clean", "").replace("_", " ").title(),
               "Revenue", "Transactions", "Unique Customers", "% of Total"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=next_row, column=i, value=h)
    _format_header_row(ws, next_row, len(headers))

    data_start = next_row + 1
    for i, row in enumerate(items):
        r = data_start + i
        ws.cell(row=r, column=1, value=row[key]).alignment = LEFT
        ws.cell(row=r, column=2, value=row["revenue"]).number_format = FMT_CURRENCY
        ws.cell(row=r, column=3, value=row["transactions"]).number_format = FMT_INT
        ws.cell(row=r, column=4, value=row["unique_customers"]).number_format = FMT_INT
        ws.cell(row=r, column=5, value=row["pct_of_total"] / 100).number_format = FMT_PERCENT

        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.font = BODY_FONT
            if i % 2 == 1:
                cell.fill = ALT_ROW_FILL

    # Chart
    chart = BarChart()
    chart.type = "bar"
    chart.style = 11
    chart.title = chart_title
    chart.y_axis.title = None
    chart.x_axis.title = "Revenue (USD)"
    chart.legend = None

    data_ref = Reference(ws, min_col=2, min_row=next_row, max_row=data_start + len(items) - 1, max_col=2)
    cats_ref = Reference(ws, min_col=1, min_row=data_start, max_row=data_start + len(items) - 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 10
    chart.width = 18

    ws.add_chart(chart, f"G{next_row}")

    for col, w in zip("ABCDE", [28, 18, 14, 18, 12]):
        ws.column_dimensions[col].width = w


def build_segment_sheet(wb: Workbook, results: dict) -> None:
    _build_breakdown_sheet(
        wb,
        title="Industry Segment Breakdown",
        sheet_name="Segments",
        subtitle="Revenue and customer count by industry vertical.",
        items=results["segment_breakdown"],
        key="industry_segment_clean",
        chart_title="Revenue by Industry Segment",
    )


def build_region_sheet(wb: Workbook, results: dict) -> None:
    _build_breakdown_sheet(
        wb,
        title="Geographic Region Breakdown",
        sheet_name="Regions",
        subtitle="Revenue and customer count by sales region.",
        items=results["region_breakdown"],
        key="region_clean",
        chart_title="Revenue by Region",
    )


def build_weekly_sheet(wb: Workbook, weekly_df: pd.DataFrame) -> None:
    """Sheet 5: weekly revenue with line chart. Trims incomplete trailing weeks."""
    ws = wb.create_sheet("Weekly Trend")

    # Flag the last row if its transaction count is much lower than the median
    # (a signal of an incomplete trailing week -- fixes the artifact Phase 3 flagged).
    # We use 40% of median: conservative enough to avoid false positives on
    # genuinely slow weeks, but catches the obvious partial-week case where
    # the week holds only 1-2 days of data.
    median_txns = weekly_df["transactions"].median()
    weekly_df = weekly_df.copy()
    weekly_df["is_partial"] = weekly_df["transactions"] < (median_txns * 0.40)

    next_row = _apply_title(
        ws, 1,
        "Weekly Revenue Trend",
        "Revenue aggregated by ISO week. Trailing partial weeks flagged in gray.",
        span_cols=5,
    )

    headers = ["Week Ending", "Revenue", "Transactions", "WoW Change", "Notes"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=next_row, column=i, value=h)
    _format_header_row(ws, next_row, len(headers))

    data_start = next_row + 1
    for i, (_, row) in enumerate(weekly_df.iterrows()):
        r = data_start + i
        ws.cell(row=r, column=1, value=row["order_date_clean"]).number_format = FMT_DATE
        ws.cell(row=r, column=2, value=row["revenue"]).number_format = FMT_CURRENCY
        ws.cell(row=r, column=3, value=row["transactions"]).number_format = FMT_INT

        # WoW change only if not partial (avoids the misleading -65% artifact)
        if pd.notna(row["wow_change_pct"]) and not row["is_partial"]:
            ws.cell(row=r, column=4, value=row["wow_change_pct"] / 100).number_format = FMT_PERCENT
        else:
            ws.cell(row=r, column=4, value="-").alignment = CENTER

        note = "Incomplete week" if row["is_partial"] else ""
        ws.cell(row=r, column=5, value=note).alignment = LEFT

        fill = ALT_ROW_FILL if i % 2 == 1 else None
        if row["is_partial"]:
            fill = PatternFill("solid", fgColor="EAEAEA")
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.font = BODY_FONT
            if fill:
                cell.fill = fill

    # Line chart (exclude partial weeks from chart visually by stopping early)
    complete_count = int((~weekly_df["is_partial"]).sum())
    chart = LineChart()
    chart.title = "Weekly Revenue"
    chart.style = 12
    chart.y_axis.title = "Revenue (USD)"
    chart.x_axis.title = "Week"
    chart.legend = None

    data_ref = Reference(ws, min_col=2, min_row=next_row,
                         max_row=data_start + complete_count - 1, max_col=2)
    cats_ref = Reference(ws, min_col=1, min_row=data_start,
                         max_row=data_start + complete_count - 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 10
    chart.width = 20

    ws.add_chart(chart, f"G{next_row}")

    for col, w in zip("ABCDE", [14, 16, 14, 14, 20]):
        ws.column_dimensions[col].width = w

    ws.freeze_panes = ws.cell(row=data_start, column=1)


def build_pareto_curve_sheet(wb: Workbook, results: dict) -> None:
    """Sheet 6: the classic 80/20 cumulative curve -- our showpiece visual."""
    ws = wb.create_sheet("Pareto Curve")

    next_row = _apply_title(
        ws, 1,
        "Pareto Curve - Revenue Concentration",
        "How quickly does cumulative revenue reach 80%? The steeper the rise, "
        "the more concentrated the business.",
        span_cols=5,
    )

    headers = ["Rank", "Customer", "Cumulative %", "80% Line"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=next_row, column=i, value=h)
    _format_header_row(ws, next_row, len(headers))

    data_start = next_row + 1
    customers = results["top_customers"]
    for i, row in enumerate(customers):
        r = data_start + i
        ws.cell(row=r, column=1, value=row["rank"]).alignment = CENTER
        ws.cell(row=r, column=2, value=row["customer_name_clean"]).alignment = LEFT
        ws.cell(row=r, column=3, value=row["cumulative_pct"] / 100).number_format = FMT_PERCENT
        ws.cell(row=r, column=4, value=0.80).number_format = FMT_PERCENT  # reference line

        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.font = BODY_FONT
            if i % 2 == 1:
                cell.fill = ALT_ROW_FILL

    # Line chart with 2 series: cumulative % and the 80% reference line
    chart = LineChart()
    chart.title = "Pareto Curve (Cumulative % by Customer Rank)"
    chart.style = 10
    chart.y_axis.title = "Cumulative Revenue %"
    chart.x_axis.title = "Customer Rank"

    cum_ref  = Reference(ws, min_col=3, min_row=next_row,
                         max_row=data_start + len(customers) - 1, max_col=3)
    line_ref = Reference(ws, min_col=4, min_row=next_row,
                         max_row=data_start + len(customers) - 1, max_col=4)
    cats_ref = Reference(ws, min_col=1, min_row=data_start,
                         max_row=data_start + len(customers) - 1)

    chart.add_data(cum_ref,  titles_from_data=True)
    chart.add_data(line_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 12
    chart.width = 22

    ws.add_chart(chart, f"F{next_row}")

    for col, w in zip("ABCD", [8, 32, 16, 14]):
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------
def generate_report(
    analysis_path: str = "data/analysis_results.json",
    weekly_path:   str = "data/weekly_revenue.csv",
    output_dir:    str = "reports",
) -> Path:
    print("Loading analysis results...")
    results = load_analysis(analysis_path)
    weekly_df = load_weekly(weekly_path)

    print("Building Executive Summary sheet...")
    wb = Workbook()
    build_exec_summary(wb, results)

    print("Building 80/20 Customers sheet...")
    build_pareto_sheet(wb, results)

    print("Building Segments sheet...")
    build_segment_sheet(wb, results)

    print("Building Regions sheet...")
    build_region_sheet(wb, results)

    print("Building Weekly Trend sheet...")
    build_weekly_sheet(wb, weekly_df)

    print("Building Pareto Curve sheet...")
    build_pareto_curve_sheet(wb, results)

    # Output path with data-date stamp
    end_date = results["dataset_summary"]["date_range_end"]
    out = Path(output_dir) / f"weekly_report_{end_date}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    print(f"\nReport saved to: {out}")
    print(f"File size: {out.stat().st_size / 1024:.1f} KB")
    print(f"\nOpen it to see 6 sheets with embedded charts.")
    return out


if __name__ == "__main__":
    generate_report()
